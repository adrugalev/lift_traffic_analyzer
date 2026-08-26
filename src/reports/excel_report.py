"""Экспорт проекта и результатов в структурированный XLSX."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models.project import Project
from src.models.results import CalculationResult, VariantResult
from src.models.simulation import SimulationResult


HEADER_FILL = PatternFill("solid", fgColor="176B87")
SECTION_FILL = PatternFill("solid", fgColor="DDEBF2")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10, color="1F2937")
THIN_BORDER = Border(
    left=Side(style="thin", color="B6C4CC"),
    right=Side(style="thin", color="B6C4CC"),
    top=Side(style="thin", color="B6C4CC"),
    bottom=Side(style="thin", color="B6C4CC"),
)


def _write_table(sheet: object, headers: list[str], rows: list[list[object]], start_row: int = 1) -> int:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for row_offset, values in enumerate(rows, start=1):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(start_row + row_offset, column, value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
    return start_row + len(rows) + 1


def _finish_sheet(sheet: object) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(42, max(10, length + 2))
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"


def build_excel_report(
    project: Project,
    analytic: CalculationResult | None = None,
    simulation: SimulationResult | None = None,
    variants: list[VariantResult] | None = None,
) -> bytes:
    """Формирует XLSX с восемью требуемыми листами."""

    workbook = Workbook()
    workbook.remove(workbook.active)

    inputs = workbook.create_sheet("Исходные данные")
    _write_table(
        inputs,
        ["Параметр", "Значение", "Источник/статус"],
        [
            ["Название проекта", project.metadata.name, "Пользователь"],
            ["Адрес", project.metadata.address, "Пользователь"],
            ["Тип здания", project.building.building_type.value, "Пользователь"],
            ["Стандарт", project.metadata.selected_standard.value, "Пользователь"],
            ["Население при 100%", project.base_population, "Таблица этажей"],
            [
                "Коэффициент заселённости",
                project.building.occupancy_percent / 100,
                "Пользователь",
            ],
            [
                "Расчётное население",
                project.population,
                "Население при 100% × коэффициент",
            ],
            [
                "Парковочные этажи",
                ", ".join(
                    str(floor.number)
                    for floor in project.floors
                    if floor.is_parking
                )
                or "Не заданы",
                "Таблица этажей",
            ],
            [
                "Доля входящего потока с паркинга",
                (
                    f"{project.scenario().parking_incoming_share:.0%}"
                    if any(floor.is_parking for floor in project.floors)
                    else "Не применяется"
                ),
                "Сценарий пассажиропотока",
            ],
            [
                "Нормативная конфигурация",
                (
                    "ГОСТ 34758-2021 активирован"
                    if analytic and analytic.calculation_basis == "GOST_34758_2021_CLAUSE_7"
                    else "Предварительный режим"
                ),
                (
                    analytic.audit.configuration_version
                    if analytic
                    else "Расчёт не приложен"
                ),
            ],
        ],
    )
    _finish_sheet(inputs)

    floors = workbook.create_sheet("Этажи")
    _write_table(
        floors,
        [
            "Этаж",
            "Отметка, м",
            "Высота, м",
            "Назначение",
            "Население при 100%",
            "Расчётное население",
            "Паркинг",
            "Основной вход",
        ],
        [
            [
                floor.number,
                floor.elevation_m,
                floor.floor_height_m,
                floor.purpose,
                floor.population,
                project.effective_floor_population(floor),
                floor.is_parking,
                floor.is_main_entrance,
            ]
            for floor in project.floors
        ],
    )
    _finish_sheet(floors)

    elevators = workbook.create_sheet("Лифты")
    _write_table(
        elevators,
        [
            "Группа",
            "Лифт",
            "Г/п, кг",
            "Номинал, пасс.",
            "Заполнение",
            "Скорость, м/с",
            "Ускорение, м/с²",
            "Замедление, м/с²",
            "Рывок, м/с³",
            "Дверь, м",
            "Тип открывания дверей",
            "Открытие, с",
            "Закрытие, с",
            "Остановки",
        ],
        [
            [
                group.name,
                elevator.name,
                elevator.capacity_kg,
                elevator.nominal_passengers,
                elevator.load_factor,
                elevator.speed_mps,
                elevator.acceleration_mps2,
                elevator.deceleration_mps2,
                elevator.jerk_mps3,
                elevator.door_width_m,
                elevator.door_opening_type.value,
                elevator.door_open_time_s,
                elevator.door_close_time_s,
                elevator.stops_count,
            ]
            for group in project.elevator_groups
            for elevator in group.elevators
        ],
    )
    for cell in elevators["E"][1:]:
        cell.number_format = "0.0%"
    _finish_sheet(elevators)

    formulas = workbook.create_sheet("Формулы")
    _write_table(
        formulas,
        ["ID", "Показатель", "Формула", "Подстановка", "Результат", "Единица", "Стандарт", "Пункт", "Статус"],
        [
            [
                trace.formula_id,
                trace.title_ru,
                trace.expression,
                trace.substituted_expression,
                trace.result,
                trace.unit,
                trace.standard,
                trace.clause or "не подтверждён",
                trace.status,
            ]
            for trace in (analytic.formulas if analytic else [])
        ],
    )
    _finish_sheet(formulas)

    analytic_sheet = workbook.create_sheet("Аналитический расчёт")
    _write_table(
        analytic_sheet,
        ["Ключ", "Показатель", "Значение", "Единица", "Метод", "Соответствие", "Цель"],
        [
            [
                metric.key,
                metric.title_ru,
                metric.value,
                metric.unit,
                metric.method,
                metric.compliance.value,
                metric.target_description,
            ]
            for metric in (analytic.metrics if analytic else [])
        ],
    )
    _finish_sheet(analytic_sheet)

    simulation_sheet = workbook.create_sheet("Симуляция")
    simulation_rows: list[list[object]] = []
    if simulation:
        simulation_rows = [
            ["Среднее время ожидания пассажира (AWT)", simulation.waiting_time.mean, "с"],
            ["Медианное время ожидания (P50)", simulation.waiting_time.median, "с"],
            ["Ожидание 80% пассажиров, не более (P80)", simulation.waiting_time.percentile_80, "с"],
            ["Ожидание 90% пассажиров, не более (P90)", simulation.waiting_time.percentile_90, "с"],
            ["Ожидание 95% пассажиров, не более (P95)", simulation.waiting_time.percentile_95, "с"],
            ["Ожидание 99% пассажиров, не более (P99)", simulation.waiting_time.percentile_99, "с"],
            ["Среднее полное время до этажа назначения (TTD)", simulation.time_to_destination.mean, "с"],
            ["Среднее время поездки в кабине", simulation.average_journey_time_s, "с"],
            ["Средняя очередь, пассажиров", simulation.average_queue_length, "пасс."],
            ["Максимальная очередь, пассажиров", simulation.maximum_queue_length, "пасс."],
            ["Перевезено пассажиров (среднее за повтор)", simulation.transported_passengers, "пасс."],
            ["Не обслужено пассажиров (среднее за повтор)", simulation.unserved_passengers, "пасс."],
        ]
    _write_table(simulation_sheet, ["Показатель", "Значение", "Единица"], simulation_rows)
    _finish_sheet(simulation_sheet)

    comparison = workbook.create_sheet("Сравнение")
    _write_table(
        comparison,
        ["Вариант", "Лифты", "Г/п, кг", "Скорость, м/с", "Интервал, с", "HC5", "AWT proxy, с", "Резерв, %", "Статус", "Оценка", "Категория"],
        [
            [
                item.variant_name,
                item.elevator_count,
                item.capacity_kg,
                item.speed_mps,
                item.interval_s,
                item.handling_capacity_5min,
                item.average_wait_s,
                item.reserve_percent,
                item.compliance.value,
                item.score,
                item.category,
            ]
            for item in (variants or [])
        ],
    )
    _finish_sheet(comparison)

    recommendations = workbook.create_sheet("Рекомендации")
    _write_table(
        recommendations,
        ["Проблема", "Показатель", "Цель", "Факт", "Решение", "Эффект", "Ограничения"],
        [
            [
                item.problem,
                item.metric,
                item.target,
                item.actual,
                item.proposed_action,
                item.expected_effect,
                item.limitations,
            ]
            for item in (analytic.recommendations if analytic else [])
        ],
    )
    _finish_sheet(recommendations)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
